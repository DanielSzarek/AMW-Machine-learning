import requests

from openpyxl import Workbook
from bs4 import BeautifulSoup


class WebScrapping:
    def __init__(self):
        self.wb = Workbook()
        self.ws_exchange = self.wb.active
        self.ws_exchange.title = "Giełda"
        self.ws_links = self.wb.create_sheet("Linki")
        self.ws_filmweb = self.wb.create_sheet("Filmweb")

    def exchange_task(self):
        pass

    def links_task(self):
        response = requests.get("https://djangopackages.org/")
        html_doc = response.text
        soup = BeautifulSoup(html_doc, 'html.parser')
        for link in soup.find_all('a'):
            try:
                self.ws_links.append([link.get('href').strip()])  # TODO sprawdzić
            except AttributeError:
                print(f"Niepoprawny znacznik <a>: {link}")

    def filmweb_task(self):
        response = requests.get("https://www.filmweb.pl/film/Nienawistna+%C3%B3semka-2015-714192")
        html_doc = response.text
        soup = BeautifulSoup(html_doc, 'html.parser')
        direction = soup.find(text="reżyseria").next_element.find("a").text.strip()
        date_of_premiere = soup.find(text="premiera").next_element.find("a").text.strip()
        rating = soup.find("span", itemprop="ratingValue").text.strip()
        boxoffice = soup.find(text="boxoffice").next_element.text
        self.ws_filmweb.append([
            "Reżyser",
            "Data premiery",
            "Boxoffice",
            "Ocena filmu"
        ])
        self.ws_filmweb.append([
            direction,
            date_of_premiere,
            boxoffice,
            rating
        ])

    def save_document(self, filename='document.xlsx'):
        self.wb.save(filename)


ws = WebScrapping()
ws.links_task()
ws.filmweb_task()
ws.save_document()
