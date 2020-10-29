import string
import requests

from re import compile
from random import choices
from openpyxl import Workbook
from bs4 import BeautifulSoup


class WebScrapping:
    def __init__(self):
        self.wb = Workbook()
        self.ws_exchange = self.wb.active
        self.ws_exchange.title = "Giełda"
        self.ws_links = self.wb.create_sheet("Linki")
        self.ws_filmweb = self.wb.create_sheet("Filmweb")

    @staticmethod
    def exchange_get_code():
        return "".join(choices(string.ascii_lowercase, k=3))

    @staticmethod
    def exchange_get_url(code):
        return f"https://stooq.pl/q/?s={code}"

    def exchange_parse(self, htmls_to_parse):
        self.ws_exchange.append([
            'Kod',
            'Kurs',
            'Zmiana',
            'Ilość transakcji'
        ])
        for code, html in htmls_to_parse.items():
            soup = BeautifulSoup(html, 'html.parser')
            price = soup.find("span", id=compile(f"aq_{code}_c\d")).text
            change = soup.find("span", id=f"aq_{code}_m3").text.strip("()")
            transactions_amount = soup.find("span", id=f"aq_{code}_n1").text
            self.ws_exchange.append([
                code,
                price,
                change,
                transactions_amount
            ])

    def exchange_task(self):
        htmls_to_parse = {}
        while len(htmls_to_parse) < 5:
            code = self.exchange_get_code()
            url = self.exchange_get_url(code)
            response = requests.get(url)
            if response.status_code == 200 and "nie istnieje w bazie" not in response.text:
                htmls_to_parse[code] = response.text
        self.exchange_parse(htmls_to_parse)

    def links_task(self):
        response = requests.get("https://djangopackages.org/")
        html_doc = response.text
        soup = BeautifulSoup(html_doc, 'html.parser')
        for link in soup.find_all('a'):
            try:
                self.ws_links.append([link.get('href').strip()])  # TODO sprawdzić
            except AttributeError:
                print(f"Niepoprawny znacznik <a>: {link}")

    def filmweb_task(self):
        response = requests.get("https://www.filmweb.pl/film/Nienawistna+%C3%B3semka-2015-714192")
        html_doc = response.text
        soup = BeautifulSoup(html_doc, 'html.parser')
        direction = soup.find(text="reżyseria").next_element.find("a").text.strip()
        date_of_premiere = soup.find(text="premiera").next_element.find("a").text.strip()
        rating = soup.find("span", itemprop="ratingValue").text.strip()
        boxoffice = soup.find(text="boxoffice").next_element.text
        self.ws_filmweb.append([
            "Reżyser",
            "Data premiery",
            "Boxoffice",
            "Ocena filmu"
        ])
        self.ws_filmweb.append([
            direction,
            date_of_premiere,
            boxoffice,
            rating
        ])

    def save_document(self, filename='Szarek-175IC_A2.xlsx'):
        self.wb.save(filename)


if __name__ == "__main__":
    ws = WebScrapping()
    ws.exchange_task()
    ws.links_task()
    ws.filmweb_task()
    ws.save_document()
